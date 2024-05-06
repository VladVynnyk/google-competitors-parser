from difflib import SequenceMatcher
import requests
import heapq
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook

import asyncio
import aiohttp
import yarl
from concurrent.futures import ThreadPoolExecutor


def is_similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

async def make_async_request_to_google_product(session, url, semaphore):
    async with semaphore:
        new_url = yarl.URL(url, encoded=True)
        header={
            "User-Agent": "python-requests/2.31.0",
            "Connection": "keep-alive",
            "Content-Type": "text/json"
            }
        async with session.get(new_url, headers=header) as r:
            if r.status == 429:
                retry_after = int(r.headers.get("Retry-After", "1"))
                print(f"Rate limit exceeded. Waiting for {retry_after} seconds before retrying.")
                await asyncio.sleep(retry_after)
                return await make_async_request_to_google_product(session, url, semaphore)
            elif r.status != 200:
                print("Status: ", r)
                r.raise_for_status()
            # elif r.status == 403:
                # print(r)
            # result = await r.json()
            print("---------------------------------")
            print(r.text)
            return r

async def make_requests_to_all_products(sesison, urls, semaphore):
    tasks = []
    for url in urls:
        task = asyncio.create_task(make_async_request_to_google_product(sesison, url, semaphore=semaphore))
        tasks.append(task)
    res = await asyncio.gather(*tasks)
    return res

async def make_requests_to_google_search(urls: list):
    rate_limit = 5
    semaphore = asyncio.Semaphore(rate_limit)
    async with aiohttp.ClientSession() as session:
        htmls = await make_requests_to_all_products(session, urls, semaphore)
    return htmls

def parse_google_ads(url: str):
    print("URL: ", url)
    print("_____________________________________: ", url)
    response = requests.get(url)
        
    if response.status_code == 200:
            html_content = response.text

            soup = BeautifulSoup(html_content, 'html.parser')

            ads_pannel = soup.find_all("div", id="bGmlqc")
            ads_name_headers = soup.html.find_all('h4', class_='fol5Z')
            ads_prices = soup.find_all('div', class_='pSNTSe')
            ads_items = soup.find_all("div", class_='pla-unit-container')
            ads_links = soup.html.find_all("a", class_="pla-unit-link")
            ads_images = soup.find_all("div", class_="Gor6zc")
            ads_sites = soup.find_all("div", class_="BZuDuc")

            product_names = []
            product_prices = []
            product_sites = []

            for item in ads_images:
                print("IMAGES: -------------------------")
                current_name = item.img['alt']
                current_name[20:]
                print(current_name[18:])
                product_names.append(current_name[18:])


            for item in ads_prices:
                current_price = item.string
                [new_price, another_value] = current_price.split(",")
                numeric_part = ''.join(char for char in new_price if char.isdigit() or char == '.') 
                product_prices.append(float(numeric_part))

            for item in ads_sites:
                product_sites.append(item.string)
            

            # dict_of_items = dict(zip(product_names, product_prices))
            # for product, price in dict_of_items.items():
            #     print(f'{product}: {price}')
                
            product_info = {}    
                
            for name, price, site in zip(product_names, product_prices, product_sites):
                if name not in product_info:
                    product_info[name] = []
                product_info[name].append([site, price])


            # STEPS: 
            # 1.Search for the cheapest products on the ads
            # 2.Write them to excel
            # 3.Search for the products on the iherb
            
            # 1.
            # Commented this section ---------------
            # workbook = load_workbook('D:\TEST OF MY JOB\EXCEL\Competitors parser\Каста ціни – копія.xlsx')
            # source_sheet = workbook.active
            # --------------------------------------


            smallest_values = heapq.nsmallest(3, product_prices)
            print("Smallest values: ", smallest_values)

            smallest_products = []
            for item in product_info:
                for y in smallest_values:
                    print(item, product_info[item])
                    print("Y:", y)
                    if y == product_info[item][0][1]:
                        # print("How similar: ", is_similar(name_of_product, item))
                        # if is_similar(name_of_product, item) >= 0.3:
                            smallest_products.append(item)

            print("--------------------------------")
            print("Cheapest products: ", smallest_products)
            print("Length of cheapest products: ", len(smallest_values))
            result = []

            for product in smallest_products:
                print("One of the cheapest products: ", product, product_info[product])
                # result.append(product_info[product])
                result.append(product_info[product][0])
                # return product_info[product]
            return result

 
# parse_google_ads("https://www.google.com/search?q=protein+whey")
# parse_google_ads("https://www.google.com/search?client=opera&q=protein+whey&sourceid=opera&ie=UTF-8&oe=UTF-8")


def main():
    workbook = load_workbook('./Testfile.xlsx')
    source_sheet = workbook.active

    products = []
    
    product_urls = [] # array with urls to google products
    for col in source_sheet.iter_cols(min_row=2, max_row=58, min_col=5, max_col=5):
        for cell in col:
            current_product_name = cell.value
            edited_product_name = current_product_name.replace(" ", "+")
            url = "https://www.google.com/search?client=opera&q="+edited_product_name+"&sourceid=opera&ie=UTF-8&oe=UTF-8"
            # url = "https://www.google.com/search?q="+edited_product_name
            product_urls.append(url)
    
    # print("URLS: ", product_urls)

    row_index = 2 
    for col in source_sheet.iter_cols(min_row=2, max_row=58, min_col=5, max_col=5):
        for cell in col:
            current_product_name = cell.value
            edited_product_name = current_product_name.replace(" ", "+")
            url = "https://www.google.com/search?client=opera&q="+edited_product_name+"&sourceid=opera&ie=UTF-8&oe=UTF-8"
            print(edited_product_name)
            print(url)

            # info = parse_google_ads(product_urls, cell.value)


            # loop = asyncio.get_event_loop()
            # loop = asyncio.new_event_loop()
            # asyncio.set_event_loop(loop)
            # google_results = loop.run_until_complete(make_requests_to_google_search(product_urls))

            # Вставка значень info в клітинки J, K, L
            # print("info:", google_results)
            # print("info:", info)
            # if info != None and len(info)>=1:
            #     source_sheet.cell(row=row_index, column=10).value = info[0][0]  # Значення info[0] у стовпець J
            #     source_sheet.cell(row=row_index, column=11).value = info[0][1]  # Значення info[0] у стовпець J
            #     print("SOURCE: ", info[0][0])
            #     print("PRICE: ", info[0][1])
            #     if len(info) == 2 or len(info) >= 2:
            #         source_sheet.cell(row=row_index, column=12).value = info[1][0]  # Значення info[1] у стовпець K
            #         source_sheet.cell(row=row_index, column=13).value = info[1][1]  # Значення info[1] у стовпець K
            #         print("SOURCE: ", info[1][0])
            #         print("PRICE: ", info[1][1])
            #     if len(info) >= 3:
            #         source_sheet.cell(row=row_index, column=14).value = info[2][0]  # Значення info[2] у стовпець L
            #         source_sheet.cell(row=row_index, column=15).value = info[2][1]  # Значення info[2] у стовпець L
            #         print("SOURCE: ", info[2][0])
            #         print("PRICE: ", info[2][1])
            #     row_index += 1
            # else:
            #     row_index += 1
            #     continue

            # row_index += 1  # Збільшення індексу рядка для наступної вставки
    MAX_THREADS = 4
    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        titles = list(executor.map(parse_google_ads, product_urls))

    print(titles)
    # workbook.save('./Testfile.xlsx')


main()
from difflib import SequenceMatcher
import requests
import heapq
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook

from urllib.parse import urlparse, parse_qs

def is_similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def extract_product_link(ad_url):
    parsed_url = urlparse(ad_url)
    query_params = parse_qs(parsed_url.query)
    if 'adurl' in query_params:
        product_link = query_params['adurl'][0]
        print("Product link: ", product_link)
        return product_link
    else:
        return None


def parse_google_ads(url: str, name_of_product: str):
    response = requests.get(url)
 
    if response.status_code == 200:
            html_content = response.text
 
            soup = BeautifulSoup(html_content, 'html.parser')
 
            ads_pannel = soup.find_all("div", id="bGmlqc")
            ads_name_headers = soup.html.find_all('h4', class_='fol5Z')
            ads_prices = soup.find_all('div', class_='pSNTSe')
            ads_items = soup.find_all("div", class_='pla-unit-container')
            ads_links = soup.find_all("a", class_="vGg33Ymfm0s__pla-unit-link")
            ads_images = soup.find_all("div", class_="Gor6zc")
            ads_sites = soup.find_all("div", class_="BZuDuc")

            

            product_names = []
            product_prices = []
            product_sites = []
            product_links = []

            print("ADS LINKS: ", ads_pannel)
            print("ADS LINKS 2: ", ads_links)
            for link in ads_links:
                print("EXTRACTED PRODUCT LINK: ", extract_product_link(link['href']))
                print("SINGLE LINK: ", link['href'])
                current_link = link['href']
                product_links.append(current_link)

            for item in ads_images:
                # print("IMAGES: -------------------------")
                current_name = item.img['alt']
                current_name[20:]
                # print(current_name[18:])
                product_names.append(current_name[18:])


            for item in ads_prices:
                current_price = item.string
                [new_price, another_value] = current_price.split(",")
                numeric_part = ''.join(char for char in new_price if char.isdigit() or char == '.') 
                product_prices.append(float(numeric_part))

            for item in ads_sites:
                product_sites.append(item.string)
            

            # dict_of_items = dict(zip(product_names, product_prices))
            # for product, price in dict_of_items.items():
            #     print(f'{product}: {price}')
                
            product_info = {}    
                
            for name, price, site, link in zip(product_names, product_prices, product_sites, product_links):
                if name not in product_info:
                    product_info[name] = []
                product_info[name].append([site, price, link])


            # STEPS: 
            # 1.Search for the cheapest products on the ads
            # 2.Write them to excel
            # 3.Search for the products on the iherb
            
            # 1.
            # Commented this section ---------------
            # workbook = load_workbook('D:\TEST OF MY JOB\EXCEL\Competitors parser\Каста ціни – копія.xlsx')
            # source_sheet = workbook.active
            # --------------------------------------


            smallest_values = heapq.nsmallest(5, product_prices) # by default 3
            # print("Smallest values: ", smallest_values)

            smallest_products = []
            for item in product_info:
                for y in smallest_values:
                    # print(item, product_info[item])
                    # print("Y:", y)
                    if y == product_info[item][0][1]:
                        # print("How similar: ", is_similar(name_of_product, item))
                        if is_similar(name_of_product, item) >= 0.3:
                            smallest_products.append(item)

            print("--------------------------------")
            print("Cheapest products: ", smallest_products)
            print("Length of cheapest products: ", len(smallest_values))
            result = []

            for product in smallest_products:
                print("One of the cheapest products: ", product, product_info[product])
                # result.append(product_info[product])
                result.append(product_info[product][0])
                # return product_info[product]
            return result

 
# parse_google_ads("https://www.google.com/search?q=protein+whey")
# parse_google_ads("https://www.google.com/search?client=opera&q=protein+whey&sourceid=opera&ie=UTF-8&oe=UTF-8")


def main():
    # workbook = load_workbook('./Testfile.xlsx')
    workbook = load_workbook('./Оборотні_профільних_ціни_конкурентів_для_HF.xlsx')
    source_sheet = workbook.active

    products = []

    # for col in source_sheet.iter_cols(min_row=2, max_row=58, min_col=5, max_col=5):
    #     for cell in col:
    #         current_product_name = cell.value
    #         edited_product_name = cell.value.replace(" ", "+")
    #         url = "https://www.google.com/search?client=opera&q="+edited_product_name+"&sourceid=opera&ie=UTF-8&oe=UTF-8"
    #         print(edited_product_name)
    #         print(url)
    #         info = parse_google_ads(url)
    #         products.append(info)
    
    row_index = 2 
    # for col in source_sheet.iter_cols(min_row=2, max_row=58, min_col=5, max_col=5):
    for col in source_sheet.iter_cols(min_row=2, max_row=242, min_col=3, max_col=3):
        for cell in col:
            current_product_name = cell.value
            edited_product_name = current_product_name.replace(" ", "+")
            url = "https://www.google.com/search?client=opera&q="+edited_product_name+"&sourceid=opera&ie=UTF-8&oe=UTF-8"
            print(edited_product_name)
            print(url)
            info = parse_google_ads(url, cell.value)

            # Вставка значень info в клітинки J, K, L
            print("info:", info)
            if info != None and len(info)>=1:
                # source_sheet.cell(row=row_index, column=10).value = info[0][0]  # Значення info[0] у стовпець J
                # source_sheet.cell(row=row_index, column=11).value = info[0][2]  # Значення info[0] у стовпець J
                # source_sheet.cell(row=row_index, column=12).value = info[0][1]  # Значення info[0] у стовпець J


                # source_sheet.cell(row=row_index, column=9).value = info[0][0]  # Значення info[0] у стовпець J
                source_sheet.cell(row=row_index, column=10).value = info[0][2]  # Значення info[0] у стовпець J
                source_sheet.cell(row=row_index, column=9).value = info[0][1]  # Значення info[0] у стовпець J
                print("SOURCE: ", info[0][0])
                print("PRICE: ", info[0][1])
                print("LINK: ", info[0][2])
                if len(info) == 2 or len(info) >= 2:
                    # source_sheet.cell(row=row_index, column=13).value = info[1][0]  # Значення info[1] у стовпець K
                    # source_sheet.cell(row=row_index, column=14).value = info[1][2]  # Значення info[1] у стовпець K
                    # source_sheet.cell(row=row_index, column=15).value = info[1][1]  # Значення info[1] у стовпець K


                    # source_sheet.cell(row=row_index, column=12).value = info[1][0]  # Значення info[1] у стовпець K
                    source_sheet.cell(row=row_index, column=13).value = info[1][2]  # Значення info[1] у стовпець K
                    source_sheet.cell(row=row_index, column=12).value = info[1][1]  # Значення info[1] у стовпець K
                    print("SOURCE: ", info[1][0])
                    print("PRICE: ", info[1][1])
                    print("LINK: ", info[1][2])
                if len(info) == 3:
                    # source_sheet.cell(row=row_index, column=16).value = info[2][0]  # Значення info[2] у стовпець L
                    # source_sheet.cell(row=row_index, column=17).value = info[2][2]  # Значення info[2] у стовпець L
                    # source_sheet.cell(row=row_index, column=18).value = info[2][1]  # Значення info[2] у стовпець L


                    # source_sheet.cell(row=row_index, column=15).value = info[2][0]  # Значення info[2] у стовпець L
                    source_sheet.cell(row=row_index, column=16).value = info[2][2]  # Значення info[2] у стовпець L
                    source_sheet.cell(row=row_index, column=15).value = info[2][1]  # Значення info[2] у стовпець L
                    print("SOURCE: ", info[2][0])
                    print("PRICE: ", info[2][1])
                    print("LINK: ", info[2][2])
                if len(info) == 4:
                    # source_sheet.cell(row=row_index, column=16).value = info[2][0]  # Значення info[2] у стовпець L
                    # source_sheet.cell(row=row_index, column=17).value = info[2][2]  # Значення info[2] у стовпець L
                    # source_sheet.cell(row=row_index, column=18).value = info[2][1]  # Значення info[2] у стовпець L


                    # source_sheet.cell(row=row_index, column=15).value = info[2][0]  # Значення info[2] у стовпець L
                    source_sheet.cell(row=row_index, column=19).value = info[2][2]  # Значення info[2] у стовпець L
                    source_sheet.cell(row=row_index, column=18).value = info[2][1]  # Значення info[2] у стовпець L
                    print("SOURCE: ", info[2][0])
                    print("PRICE: ", info[2][1])
                    print("LINK: ", info[2][2])
                if len(info) == 5:
                    # source_sheet.cell(row=row_index, column=16).value = info[2][0]  # Значення info[2] у стовпець L
                    # source_sheet.cell(row=row_index, column=17).value = info[2][2]  # Значення info[2] у стовпець L
                    # source_sheet.cell(row=row_index, column=18).value = info[2][1]  # Значення info[2] у стовпець L


                    # source_sheet.cell(row=row_index, column=15).value = info[2][0]  # Значення info[2] у стовпець L
                    source_sheet.cell(row=row_index, column=22).value = info[2][2]  # Значення info[2] у стовпець L
                    source_sheet.cell(row=row_index, column=21).value = info[2][1]  # Значення info[2] у стовпець L
                    print("SOURCE: ", info[2][0])
                    print("PRICE: ", info[2][1])
                    print("LINK: ", info[2][2])
                row_index += 1
            else:
                row_index += 1
                continue

            # row_index += 1  # Збільшення індексу рядка для наступної вставки

    # workbook.save('./Testfile.xlsx')
    workbook.save('./Оборотні_профільних_ціни_конкурентів_для_HF.xlsx')


main()
import time

from difflib import SequenceMatcher
import requests
import heapq
from bs4 import BeautifulSoup
from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl import Workbook

from openai import OpenAI

from urllib.parse import urlparse, parse_qs

from Parser import Parser
from utils import get_random_user_agent, format_price, create_messages_for_ai, is_similar, extract_product_link


config = dotenv_values(".env")
print("config " , config)
CHATGPT_API_KEY = config["CHATGPT_API_KEY"]
 
def main():
    parser = Parser()
    client = OpenAI(api_key=CHATGPT_API_KEY)

    workbook = load_workbook('./example (3) 99.xlsx')
    source_sheet = workbook.active

    max_row = source_sheet.max_row
    
    products = []

    
    row_index = 2
    for col in source_sheet.iter_cols(min_row=2, max_row=max_row, min_col=3, max_col=3):
        for cell in col:
            current_product_name = cell.value
            edited_product_name = current_product_name.replace(" ", "+")
            url = "https://www.google.com/search?q="+edited_product_name
            print(edited_product_name)
            print(url)
            info = parser.parse_google_ads(url, cell.value, client_of_ai=client, isIherb=False)
            # {product1: {source: Bigl.ua, price:1000, link:https://}, product2: {source: Bigl.ua, price:1000, link:https://}}

            print("info:", info)
            # print("info:", len(info))
            # time.sleep(10)

            if info is not None and len(info) >= 1:
                column_index = 4  # Starting column index
                for key, value in info.items():
                    if column_index >= 15:
                        column_index = 4 
                        continue
                    else:
                        source_sheet.cell(row=row_index, column=column_index).value = value['site']
                        source_sheet.cell(row=row_index, column=column_index + 1).value = value['link']
                        source_sheet.cell(row=row_index, column=column_index + 2).value = value['price']
                        print("PRICE:", value['price'])
                        print("LINK:", value['link'])
                        column_index += 3  # Move to the next pair of columns

                    workbook.save('./example (3) 99 - v2.2.xlsx')
                    time.sleep(10)

                row_index += 1
            else:
                row_index += 1
                continue


    row_index = 2
    for col in source_sheet.iter_cols(min_row=2, max_row=max_row, min_col=3, max_col=3):
        for cell in col:
            current_product_name = cell.value
            edited_product_name = current_product_name.replace(" ", "+")
            url = "https://www.google.com/search?q="+edited_product_name+"+iherb"
            print(edited_product_name)
            print(url)
            info = parser.parse_google_ads(url, cell.value, client_of_ai=client, isIherb=True)
            # {product1: {source: Bigl.ua, price:1000, link:https://}, product2: {source: Bigl.ua, price:1000, link:https://}}

            print("info:", info)
            # print("info:", len(info))
            # time.sleep(10)

            if info is not None and len(info) >= 1:
                column_index = 16  # Starting column index (for iherb must be bigger than for google)
                for key, value in info.items():
                    # if column_index >= 15:
                        # column_index = 4 
                        # continue
                    # else:
                        source_sheet.cell(row=row_index, column=column_index).value = value['site']
                        source_sheet.cell(row=row_index, column=column_index + 1).value = value['link']
                        source_sheet.cell(row=row_index, column=column_index + 2).value = value['price']
                        print("PRICE:", value['price'])
                        print("LINK:", value['link'])
                        column_index += 3  # Move to the next pair of columns

                        workbook.save('./example (3) 99 - v2.3.xlsx')
                        time.sleep(10)
                row_index += 1
            else:
                row_index += 1
                continue





    # workbook.save('./Testfile.xlsx')
    # workbook.save('./Оборотні_профільних_ціни_конкурентів_для_HF.xlsx')


main()